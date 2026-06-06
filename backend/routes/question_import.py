"""Question Import & Option Completion Tool — Routes (Phase 2)."""
import os as _os
import json as _json
import uuid as _uuid
import re as _re
import tempfile as _tempfile
import aiofiles as _aiofiles
import io as _io
from pathlib import Path as _Path
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from typing import List, Optional
from database import db, logger
from models import (
    ImportJob, ImportFileInfo, ParsedQuestion, ImportResponse,
    ValidationSummary, ValidationResult, GenerateOptionsResponse, ExportQuestion
)
from auth import get_current_user, get_admin_user
from services.import_job import (
    create_import_job, get_import_job, update_import_job,
    add_file_to_job, update_file_status, add_questions_to_job,
    set_job_status, validate_import_job, update_question_generated_options,
    get_questions_for_export
)
from services.ocr_service import extract_text_from_pdf, extract_text_from_markdown
from services.question_parser import parse_questions_from_text, parse_questions_from_text_async
from services.option_generator import generate_for_questions

router = APIRouter(prefix="/api", tags=["question-import"])

ALLOWED_EXTENSIONS = {".pdf", ".md", ".markdown", ".txt"}
UPLOAD_DIR = _Path(_os.environ.get("IMPORT_UPLOAD_DIR", _tempfile.gettempdir())) / "question_import"


@router.post("/admin/question-import", response_model=ImportResponse)
async def upload_import_files(
    files: List[UploadFile] = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload PDF and/or Markdown files for question import. Creates an ImportJob and returns its ID."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    # Validate file types
    for f in files:
        ext = _os.path.splitext(f.filename or "")[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type '{ext}' for '{f.filename}'. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
            )

    # Create import job
    job = await create_import_job()

    # Save files to disk and register in job
    job_dir = UPLOAD_DIR / job.id
    job_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        file_type = "pdf" if f.filename and f.filename.lower().endswith(".pdf") else "markdown"
        save_path = job_dir / (f.filename or "unnamed")
        content = await f.read()
        _aiofiles;  # ensure imported
        async with _aiofiles.open(str(save_path), "wb") as out:
            await out.write(content)

        file_info = ImportFileInfo(
            filename=f.filename or "unnamed",
            file_type=file_type,
            status="uploaded"
        )
        await add_file_to_job(job.id, file_info)

    return ImportResponse(import_id=job.id, status="uploaded")


async def _background_process_job(import_id: str):
    """Run extraction in the background. Updates job status in MongoDB as it progresses."""
    try:
        job = await get_import_job(import_id)
        if not job:
            logger.error(f"[BACKGROUND] Job {import_id} not found — already deleted?")
            return

        job_dir = UPLOAD_DIR / job.id
        all_questions = []
        errors = []

        for file_info in job.files:
            file_path = job_dir / file_info.filename
            if not file_path.exists():
                await update_file_status(import_id, file_info.filename, "failed", error="File not found on disk")
                errors.append({"filename": file_info.filename, "error": "File not found"})
                continue

            try:
                await update_file_status(import_id, file_info.filename, "processing")

                if file_info.file_type == "pdf":
                    text = extract_text_from_pdf(str(file_path))
                else:
                    text = extract_text_from_markdown(str(file_path))

                if not text or not text.strip():
                    raise ValueError("Extracted text is empty")

                parsed = await parse_questions_from_text_async(text, source_file=file_info.filename)
                if parsed:
                    # Run recovery on extracted questions using original file text
                    recovered_count = 0
                    if text and parsed:
                        import re as _re
                        sections = _re.split(r"(?=^## \d+\.)", text, flags=_re.MULTILINE)
                        logger.info(f"[BACKGROUND] Recovery: {len(parsed)} questions, {len(sections)} source sections")
                        for pq in parsed:
                            needs_opts = not pq.options or len(pq.options) < 2
                            needs_ans = not pq.correct_answers
                            if not needs_opts and not needs_ans:
                                continue
                            qtext = pq.question.strip()
                            best_sec = None
                            # Match by number
                            nm = _re.match(r"^\s*(\d+)\s*[.)]\s*", qtext)
                            if nm:
                                for sec in sections:
                                    sm = _re.match(r"^##\s*(\d+)\.\s*", sec.strip())
                                    if sm and sm.group(1) == nm.group(1):
                                        best_sec = sec
                                        break
                            # Fall back to word overlap
                            if not best_sec:
                                qwords = set(_re.sub(r"^\d+\.\s*", "", qtext.lower())[:100].split())
                                best_score = 0
                                for sec in sections:
                                    sec_words = set(_re.sub(r"^## \d+\.\s*", "", sec.strip().lower())[:100].split())
                                    score = len(qwords & sec_words)
                                    if score > best_score:
                                        best_score = score
                                        best_sec = sec
                                if best_score < 2:
                                    continue
                            # Extract bullet options
                            if needs_opts:
                                opt_lines = _re.findall(r"^-\s+(.*)", best_sec, _re.MULTILINE)
                                clean_opts = [o.strip() for o in opt_lines if o.strip() and o.strip() != "-"]
                                clean_opts = list(dict.fromkeys(clean_opts))  # dedup preserving order
                                if len(clean_opts) >= 2:
                                    pq.options = clean_opts
                                    if pq.type in ("", "unknown"):
                                        pq.type = "single"
                            # Extract answer
                            if needs_ans or (pq.options and not pq.correct_answers):
                                am = _re.search(r"\*\*Answer:\*\*\s*(.*?)(?:\n_(?:Page|Seite)|$)", best_sec, _re.IGNORECASE | _re.DOTALL)
                                if am:
                                    ans_text = am.group(1).strip()
                                    if pq.options:
                                        matched = False
                                        for opt in pq.options:
                                            optc = _re.sub(r"^[A-Za-z][)\s.]+\s*", "", opt).strip().lower()
                                            ansl = ans_text.strip().lower()
                                            if ansl == optc or ansl in optc or optc in ansl:
                                                pq.correct_answers = [opt]
                                                matched = True
                                                break
                                        if not matched:
                                            pq.correct_answers = [ans_text]
                                    else:
                                        pq.correct_answers = [ans_text]
                                    recovered_count += 1
                            # Also count option-only recoveries
                            if needs_opts and pq.options and len(pq.options) >= 2:
                                recovered_count += 1
                        if recovered_count:
                            logger.info(f"[BACKGROUND] Recovery: patched {recovered_count} questions for {file_info.filename}")
                    all_questions.extend(parsed)
                    await update_file_status(import_id, file_info.filename, "parsed", questions_count=len(parsed))
                else:
                    fallback = ParsedQuestion(
                        question=f"[RAW TEXT from {file_info.filename}]",
                        options=[],
                        correct_answers=[],
                        source_file=file_info.filename,
                        status="parsed",
                        error="No structured questions detected"
                    )
                    all_questions.append(fallback)
                    await update_file_status(import_id, file_info.filename, "parsed", questions_count=1)

            except Exception as e:
                logger.error(f"[BACKGROUND] Failed to process {file_info.filename}: {e}")
                await update_file_status(import_id, file_info.filename, "failed", error=str(e))
                errors.append({"filename": file_info.filename, "error": str(e)})

        if all_questions:
            await add_questions_to_job(import_id, all_questions)

        new_status = "parsed" if not errors else ("failed" if not all_questions else "parsed")
        await set_job_status(import_id, new_status)

        logger.info(f"[BACKGROUND] Job {import_id} complete: {len(all_questions)} questions, status={new_status}")

    except Exception as e:
        logger.error(f"[BACKGROUND] Job {import_id} crashed: {e}")
        try:
            await set_job_status(import_id, "failed")
        except Exception:
            pass


@router.post("/admin/question-import/{import_id}/process")
async def process_import_job(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Start processing uploaded files in the background. Returns immediately."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.status != "uploaded":
        raise HTTPException(status_code=400, detail=f"Job already processed (status: {job.status})")

    await set_job_status(import_id, "processing")
    import asyncio
    asyncio.create_task(_background_process_job(import_id))

    return {
        "import_id": import_id,
        "status": "processing",
        "message": "Extraction started in background. Poll GET /admin/question-import/{id} for status."
    }


@router.post("/admin/question-import/{import_id}/generate-options", response_model=GenerateOptionsResponse)
async def generate_question_options(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Generate AI distractors for parsed questions. Processes up to 50 questions per batch."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.status not in ("parsed", "completed"):
        raise HTTPException(status_code=400, detail=f"Job must be parsed first (status: {job.status})")

    if not job.questions:
        return GenerateOptionsResponse(import_id=import_id, processed=0, updated=0, skipped=0, failed=0, total=0)

    await set_job_status(import_id, "processing")
    result = await generate_for_questions(job.questions)

    # Persist generated options and status to DB
    for r in result["results"]:
        q_status = "completed" if r["action"] == "updated" else r["action"]
        if r.get("generated"):
            await update_question_generated_options(import_id, r["index"], r["generated"], status=q_status)
        elif r["action"] in ("failed_generation", "failed"):
            await update_question_generated_options(import_id, r["index"], [], status=q_status)

    has_failed = any(r["action"] in ("failed_generation", "failed") for r in result["results"])
    new_status = "completed" if not has_failed else ("completed_with_errors" if result["updated"] > 0 else "failed")
    await set_job_status(import_id, new_status)

    return GenerateOptionsResponse(
        import_id=import_id,
        processed=result["processed"],
        updated=result["updated"],
        skipped=result["skipped"],
        failed=result["failed"],
        total=result["total"],
        results=result["results"]
    )


@router.get("/admin/question-import/{import_id}/export/json")
async def export_import_json(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Export parsed questions with generated options as JSON."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    items = await get_questions_for_export(import_id)
    if not items:
        raise HTTPException(status_code=404, detail="Import job not found or empty")
    return {"import_id": import_id, "total": len(items), "questions": items}


@router.get("/admin/question-import/{import_id}/export/xlsx")
async def export_import_xlsx(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Export parsed questions with generated options as Excel (.xlsx)."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    items = await get_questions_for_export(import_id)
    if not items:
        raise HTTPException(status_code=404, detail="Import job not found or empty")

    try:
        import openpyxl as _xlsx
        from openpyxl.styles import Font as _Font, Alignment as _Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = _xlsx.Workbook()
    ws = wb.active
    ws.title = "Questions"
    bold = _Font(bold=True)

    headers = ["#", "Question", "Correct Answer(s)", "All Options", "Original Options", "Generated Options"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for i, q in enumerate(items, 2):
        ws.cell(row=i, column=1, value=q["index"])
        ws.cell(row=i, column=2, value=q["question"]).alignment = _Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value="; ".join(q["correct_answers"]))
        ws.cell(row=i, column=4, value="; ".join(q["final_options"]))
        ws.cell(row=i, column=5, value="; ".join(q["original_options"]))
        ws.cell(row=i, column=6, value="; ".join(q["generated_options"]))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=questions_{import_id}.xlsx"}
    )


@router.get("/admin/question-import/{import_id}/export/markdown")
async def export_import_markdown(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Export parsed questions as re-importable Markdown."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")

    from services.export_service import build_export_items, export_to_markdown
    items = build_export_items(job.questions)
    if not items:
        raise HTTPException(status_code=404, detail="No questions to export")

    metadata = {
        "import_id": import_id,
        "source": job.files[0].filename if job.files else "unknown",
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "total": len(items),
    }
    md = export_to_markdown(items, source_metadata=metadata)

    return StreamingResponse(
        iter([md]),
        media_type="text/markdown",
        headers={"Content-Disposition": f"attachment; filename=questions_{import_id}.md"}
    )


@router.post("/admin/question-import/re-import")
async def re_import_exported(
    request: Request,
    user: dict = Depends(get_current_user)
):
    """Re-import a previously exported file. Tests roundtrip fidelity."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    body = await request.json()
    text = body.get("text", "")
    format_type = body.get("format", "markdown")

    if not text:
        raise HTTPException(status_code=400, detail="No content to re-import")

    # Parse using AI extractor
    parsed = await parse_questions_from_text_async(text, source_file="re-import")

    # Build comparison report
    expected_count = body.get("expected_count")
    report = {
        "re_imported": len(parsed),
        "expected": expected_count or len(parsed),
        "match": len(parsed) == (expected_count or len(parsed)),
        "questions": [
            {
                "index": i,
                "question": q.question[:100],
                "options_count": len(q.options),
                "correct_count": len(q.correct_answers),
            }
            for i, q in enumerate(parsed)
        ],
    }

    return report


@router.get("/admin/question-import/{import_id}")
async def get_import_job_details(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Get details of an import job including parsed questions."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    return job


@router.post("/admin/question-import/{import_id}/validate")
async def validate_import_job_endpoint(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Validate all questions in an import job. Returns summary of valid/invalid questions."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")

    summary = validate_import_job(job)
    return summary


@router.post("/admin/question-import/extract-report")
async def extract_report_endpoint(
    request: Request,
    user: dict = Depends(get_current_user)
):
    """Run AI extraction on raw text and return detailed diagnostic report."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    body = await request.json()
    text = body.get("text", "")
    if not text:
        raise HTTPException(status_code=400, detail="No text provided")
    from services.question_parser import extract_with_report
    report = await extract_with_report(text, source_file="api-report")
    return report


@router.post("/admin/question-import/{import_id}/re-extract")
async def re_extract_job(
    import_id: str,
    user: dict = Depends(get_current_user)
):
    """Re-extract questions using AI extraction (overwrites existing questions)."""
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    job = await get_import_job(import_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")

    # Re-read files from disk and re-extract with AI
    job_dir = UPLOAD_DIR / job.id
    all_questions = []
    errors = []

    for file_info in job.files:
        file_path = job_dir / file_info.filename
        if not file_path.exists():
            errors.append({"filename": file_info.filename, "error": "File not found on disk"})
            continue
        try:
            if file_info.file_type == "pdf":
                from services.ocr_service import extract_text_from_pdf
                text = extract_text_from_pdf(str(file_path))
            else:
                from services.ocr_service import extract_text_from_markdown
                text = extract_text_from_markdown(str(file_path))

            parsed = await parse_questions_from_text_async(text, source_file=file_info.filename)
            if parsed:
                all_questions.extend(parsed)
        except Exception as e:
            logger.error(f"Re-extract failed for {file_info.filename}: {e}")
            errors.append({"filename": file_info.filename, "error": str(e)})

    # Replace questions in job
    if all_questions:
        await db.import_jobs.update_one(
            {"id": import_id},
            {"$set": {"questions": [q.model_dump() for q in all_questions], "status": "parsed"}}
        )

    return {
        "import_id": import_id,
        "questions_extracted": len(all_questions),
        "errors": errors,
    }
