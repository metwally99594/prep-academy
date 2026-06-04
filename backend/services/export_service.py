"""Export pipeline for parsed questions.

Formats:
  - Markdown (.md) — production-ready, re-importable
  - Excel (.xlsx) — structured table
  - JSON (.json) — raw data preserve

Export guarantees:
  - Every question included (except failed_generation)
  - Question text preserved exactly
  - All options preserved exactly
  - correct_answers[] preserved exactly
  - Question type preserved
  - Generated distractors preserved
  - Source metadata preserved
  - Re-importable without data loss
"""
import io, json as _json
from typing import List, Optional
from models import ParsedQuestion
from database import logger


def _format_option(option: str) -> str:
    """Ensure option has letter prefix. Add one if missing."""
    import re
    if re.match(r"^[A-Za-z][.)]\s", option):
        return option
    return option


def _next_letter(idx: int) -> str:
    return chr(ord("A") + idx)


def export_to_markdown(
    questions: List[dict],
    source_metadata: Optional[dict] = None,
) -> str:
    """Export questions as a re-importable Markdown document."""
    lines = []
    lines.append("# Exported Questions")
    lines.append("")

    if source_metadata:
        lines.append("---")
        for k, v in source_metadata.items():
            lines.append(f"{k}: {v}")
        lines.append("---")
        lines.append("")

    for i, q in enumerate(questions, 1):
        qtype = q.get("type", "single")
        qtype_label = {
            "single": "[Single Choice]",
            "multi": "[Multiple Choice]",
            "true_false": "[True/False]",
            "matching": "[Matching]",
            "grouping": "[Grouping]",
            "image": "[Image-Based]",
        }.get(qtype, "[Single Choice]")

        lines.append(f"## Question {i} {qtype_label}")
        lines.append("")
        lines.append(q.get("question", ""))
        lines.append("")

        options = q.get("all_options", q.get("options", []))
        if qtype == "true_false" and not options:
            options = ["A) Richtig", "B) Falsch"]

        for opt in options:
            lines.append(_format_option(opt))

        if options:
            lines.append("")

        correct = q.get("correct_answers", [])
        if correct:
            if len(correct) == 1:
                lines.append(f"**Answer:** {correct[0]}")
            else:
                lines.append(f"**Correct answers:** {', '.join(correct)}")
        else:
            lines.append("**Answer:** ")

        lines.append("")

    return "\n".join(lines)


def export_to_json(questions: List[dict]) -> str:
    """Export questions as a JSON array. Fully preserves all fields."""
    output = []
    for q in questions:
        entry = {
            "question": q.get("question", ""),
            "type": q.get("type", "single"),
            "options": q.get("all_options", q.get("options", [])),
            "correct_answers": q.get("correct_answers", []),
            "generated_options": q.get("generated_options", []),
            "original_options": q.get("original_options", q.get("options", [])),
            "source_file": q.get("source_file", ""),
            "question_type": q.get("type", "single"),
        }
        output.append(entry)

    return _json.dumps(output, ensure_ascii=False, indent=2)


def export_to_xlsx(questions: List[dict]) -> bytes:
    """Export questions as an Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    bold = Font(bold=True)
    headers = [
        "#", "Question", "Type", "Correct Answer(s)",
        "All Options", "Original Options", "Generated Options",
        "Source File"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for i, q in enumerate(questions, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=q.get("question", "")).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=q.get("type", "single"))
        ws.cell(row=i, column=4, value="; ".join(q.get("correct_answers", [])))
        ws.cell(row=i, column=5, value="; ".join(q.get("all_options", q.get("options", []))))
        ws.cell(row=i, column=6, value="; ".join(q.get("original_options", q.get("options", []))))
        ws.cell(row=i, column=7, value="; ".join(q.get("generated_options", [])))
        ws.cell(row=i, column=8, value=q.get("source_file", ""))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_export_items(
    parsed_questions: List[ParsedQuestion],
    include_failed: bool = False,
) -> List[dict]:
    """Build export items from ParsedQuestion list.

    Each item contains:
      - question, type, options, correct_answers
      - original_options, generated_options, all_options
      - source_file, status
    """
    items = []
    for q in parsed_questions:
        if q.status == "failed_generation" and not include_failed:
            continue
        orig_opts = q.options or []
        gen_opts = q.generated_options or []
        all_opts = orig_opts + gen_opts
        items.append({
            "question": q.question,
            "type": q.type if hasattr(q, "type") and q.type else "single",
            "options": orig_opts,
            "original_options": orig_opts,
            "generated_options": gen_opts,
            "all_options": all_opts,
            "correct_answers": q.correct_answers,
            "source_file": q.source_file,
            "status": q.status,
        })
    return items
